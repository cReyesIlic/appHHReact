import re
import sqlite3
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.core.config import settings


@dataclass
class HHRow:
    codigo: str
    file_path: str
    workbook_name: str
    sheet_name: str
    row_number: int
    item: str
    deliverable: str
    activity: str
    discipline: str
    role: str
    hours: float | None
    rate: float | None
    amount: float | None
    raw_text: str
    confidence: float


class HHExcelExtractor:
    def __init__(self) -> None:
        self._ensure_tables()

    def extract_file(self, codigo: str, path: str | Path) -> dict:
        file_path = Path(path)
        if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
            return {"codigo": codigo, "status": "unsupported", "rows": 0, "error": f"Formato no soportado: {file_path.suffix}"}
        rows: list[HHRow] = []
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                rows.extend(self._extract_sheet(codigo, file_path, sheet))
        except Exception as exc:
            return {"codigo": codigo, "status": "error", "rows": 0, "error": str(exc)}
        self.replace_file(codigo, str(file_path), rows)
        return {"codigo": codigo, "status": "ok", "rows": len(rows), "error": ""}

    def replace_file(self, codigo: str, file_path: str, rows: list[HHRow]) -> None:
        with sqlite3.connect(settings.sqlite_path) as conn:
            conn.execute("delete from hh_estimate_rows where codigo = ? and file_path = ?", (codigo, file_path))
            for row in rows:
                conn.execute(
                    """
                    insert into hh_estimate_rows
                    (codigo, file_path, workbook_name, sheet_name, row_number, item, deliverable, activity,
                     discipline, role, hours, rate, amount, raw_text, confidence)
                    values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        row.codigo,
                        row.file_path,
                        row.workbook_name,
                        row.sheet_name,
                        row.row_number,
                        row.item,
                        row.deliverable,
                        row.activity,
                        row.discipline,
                        row.role,
                        row.hours,
                        row.rate,
                        row.amount,
                        row.raw_text,
                        row.confidence,
                    ),
                )

    def query(self, codigo: str | None = None, text: str | None = None, limit: int = 100) -> list[dict]:
        sql = """
            select codigo, workbook_name, sheet_name, row_number, item, deliverable, activity,
                   discipline, role, hours, rate, amount, raw_text, confidence
            from hh_estimate_rows
        """
        params: list[str] = []
        where = []
        if codigo:
            where.append("codigo = ?")
            params.append(codigo.upper())
        if text:
            where.append("(raw_text like ? or deliverable like ? or activity like ?)")
            value = f"%{text}%"
            params.extend([value, value, value])
        if where:
            sql += " where " + " and ".join(where)
        sql += " order by codigo, workbook_name, sheet_name, row_number limit ?"
        params.append(str(limit))
        with sqlite3.connect(settings.sqlite_path) as conn:
            conn.row_factory = sqlite3.Row
            return [dict(row) for row in conn.execute(sql, params).fetchall()]

    def summary(self, codigo: str | None = None) -> dict:
        params: list[str] = []
        where = ""
        if codigo:
            where = "where codigo = ?"
            params.append(codigo.upper())
        with sqlite3.connect(settings.sqlite_path) as conn:
            row = conn.execute(
                f"""
                select count(*) rows,
                       count(distinct codigo) proposals,
                       count(distinct file_path) files,
                       sum(coalesce(hours, 0)) total_hours,
                       sum(coalesce(amount, 0)) total_amount,
                       sum(case when confidence >= 0.65 and hours > 0 and hours <= 20000 then hours else 0 end) total_hours_plausible,
                       sum(case when confidence >= 0.65 and amount > 0 and amount <= 100000000 then amount else 0 end) total_amount_plausible
                from hh_estimate_rows
                {where}
                """,
                params,
            ).fetchone()
        return {
            "rows": row[0],
            "proposal_count": row[1],
            "file_count": row[2],
            "total_hours": row[3],
            "total_amount": row[4],
            "total_hours_plausible": row[5],
            "total_amount_plausible": row[6],
            "quality_note": "Totales brutos son exploratorios; usar filas y totales plausibles como apoyo, validar formato del Excel antes de decision comercial.",
        }

    def _extract_sheet(self, codigo: str, file_path: Path, sheet) -> list[HHRow]:
        rows = []
        header: dict[str, int] = {}
        sheet_name_norm = self._norm(sheet.title)
        relevant_sheet = any(token in sheet_name_norm for token in ["hh", "estimacion", "horas", "tarifa", "eco", "cotizacion", "entreg"])

        for row_number, values in enumerate(sheet.iter_rows(min_row=1, max_row=600, max_col=80, values_only=True), start=1):
            cells = [self._clean(value) for value in values]
            raw_text = " | ".join(cell for cell in cells if cell)
            if not raw_text:
                continue
            maybe_header = self._header_map(cells)
            if len(maybe_header) >= 2:
                header = maybe_header
                continue
            if not relevant_sheet and not header:
                continue
            parsed = self._parse_data_row(codigo, file_path, sheet.title, row_number, cells, raw_text, header)
            if parsed:
                rows.append(parsed)
        return rows

    def _parse_data_row(
        self,
        codigo: str,
        file_path: Path,
        sheet_name: str,
        row_number: int,
        cells: list[str],
        raw_text: str,
        header: dict[str, int],
    ) -> HHRow | None:
        if self._is_rejected_summary_row(raw_text):
            return None
        numbers = [(index, self._number(value)) for index, value in enumerate(cells)]
        numbers = [(index, value) for index, value in numbers if value is not None]
        text_cells = [(index, value) for index, value in enumerate(cells) if value and self._number(value) is None]
        if not text_cells or not numbers:
            return None
        if len(numbers) == 1 and numbers[0][0] <= 1 and self._first_code(cells):
            return None

        unit_values = self._values_after_hh_unit(cells)
        hours = unit_values.get("hours") or self._value_by_header(cells, header, ["hh", "horas", "h/h", "total hh"])
        rate = unit_values.get("rate") or self._value_by_header(cells, header, ["tarifa", "valor hora", "uf/hh", "clp/hh"])
        amount = unit_values.get("amount") or self._value_by_header(cells, header, ["monto", "subtotal", "valor", "precio"])
        if hours is None:
            hours = self._best_hours(numbers)
        if amount is None and (rate is not None or self._norm(raw_text).find("costo") >= 0):
            amount = self._best_amount(numbers, hours, rate)

        deliverable = self._text_by_header(cells, header, ["entregable", "producto", "documento"])
        activity = self._text_by_header(cells, header, ["actividad", "descripcion", "alcance", "tarea"])
        discipline = self._text_by_header(cells, header, ["disciplina", "area", "especialidad"])
        role = self._text_by_header(cells, header, ["cargo", "recurso", "profesional", "categoria"])

        if not any([deliverable, activity, discipline, role]):
            deliverable = self._best_label(text_cells)
        if hours is None and amount is None:
            return None

        confidence = 0.45
        if header:
            confidence += 0.25
        if hours is not None:
            confidence += 0.2
        if deliverable or activity:
            confidence += 0.1

        return HHRow(
            codigo=codigo,
            file_path=str(file_path),
            workbook_name=file_path.name,
            sheet_name=sheet_name,
            row_number=row_number,
            item=self._first_code(cells),
            deliverable=deliverable,
            activity=activity,
            discipline=discipline,
            role=role,
            hours=hours,
            rate=rate,
            amount=amount,
            raw_text=raw_text[:1500],
            confidence=min(confidence, 1.0),
        )

    def _header_map(self, cells: list[str]) -> dict[str, int]:
        header = {}
        for index, cell in enumerate(cells):
            norm = self._norm(cell)
            if not norm:
                continue
            for token in [
                "entregable",
                "producto",
                "documento",
                "actividad",
                "descripcion",
                "alcance",
                "tarea",
                "disciplina",
                "especialidad",
                "area",
                "cargo",
                "recurso",
                "profesional",
                "categoria",
                "hh",
                "horas",
                "h/h",
                "total hh",
                "tarifa",
                "valor hora",
                "monto",
                "total",
                "subtotal",
                "precio",
            ]:
                if token in norm and token not in header:
                    header[token] = index
        return header

    def _value_by_header(self, cells: list[str], header: dict[str, int], keys: list[str]) -> float | None:
        for key in keys:
            index = header.get(key)
            if index is not None and index < len(cells):
                value = self._number(cells[index])
                if value is not None:
                    return value
        return None

    def _text_by_header(self, cells: list[str], header: dict[str, int], keys: list[str]) -> str:
        for key in keys:
            index = header.get(key)
            if index is not None and index < len(cells):
                value = cells[index].strip()
                if value and self._number(value) is None:
                    return value[:500]
        return ""

    def _best_hours(self, numbers: list[tuple[int, float]]) -> float | None:
        candidates = [value for index, value in numbers if index > 0 and 0 < value <= 100000]
        return candidates[-1] if candidates else None

    def _best_amount(self, numbers: list[tuple[int, float]], hours: float | None, rate: float | None) -> float | None:
        if hours and rate:
            return hours * rate
        candidates = [value for index, value in numbers if index > 0 and abs(value) > 1000 and value != hours]
        return candidates[-1] if candidates else None

    def _values_after_hh_unit(self, cells: list[str]) -> dict[str, float | None]:
        for index, cell in enumerate(cells):
            if self._norm(cell) in {"hh", "h/h", "horas"}:
                nums = [self._number(value) for value in cells[index + 1 : index + 4]]
                nums = [value for value in nums if value is not None]
                return {
                    "hours": nums[0] if len(nums) >= 1 else None,
                    "rate": nums[1] if len(nums) >= 2 else None,
                    "amount": nums[2] if len(nums) >= 3 else None,
                }
        return {"hours": None, "rate": None, "amount": None}

    def _is_rejected_summary_row(self, raw_text: str) -> bool:
        norm = self._norm(raw_text)
        rejected = ["total (uf)", "total uf", "valor total", "precio total", "resumen"]
        return any(token in norm for token in rejected)

    def _best_label(self, text_cells: list[tuple[int, str]]) -> str:
        sorted_cells = sorted(text_cells, key=lambda item: len(item[1]), reverse=True)
        return sorted_cells[0][1][:500] if sorted_cells else ""

    def _first_code(self, cells: list[str]) -> str:
        for cell in cells[:5]:
            if re.match(r"^[A-Za-z]?\d+(\.\d+)*$", cell.strip()):
                return cell.strip()
        return ""

    def _clean(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _number(self, value) -> float | None:
        text = str(value or "").strip()
        if not text:
            return None
        text = text.replace("\xa0", "").replace(" ", "")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", ".")
        text = re.sub(r"[^0-9.\-]", "", text)
        if text in {"", "-", ".", "-."}:
            return None
        try:
            return float(text)
        except ValueError:
            return None

    def _ensure_tables(self) -> None:
        settings.sqlite_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(settings.sqlite_path) as conn:
            conn.execute(
                """
                create table if not exists hh_estimate_rows (
                    id integer primary key autoincrement,
                    codigo text not null,
                    file_path text not null,
                    workbook_name text not null,
                    sheet_name text not null,
                    row_number integer not null,
                    item text,
                    deliverable text,
                    activity text,
                    discipline text,
                    role text,
                    hours real,
                    rate real,
                    amount real,
                    raw_text text,
                    confidence real not null
                )
                """
            )
            conn.execute("create index if not exists idx_hh_estimate_codigo on hh_estimate_rows(codigo)")
            conn.execute("create index if not exists idx_hh_estimate_file on hh_estimate_rows(file_path)")

    def _norm(self, value: object) -> str:
        text = unicodedata.normalize("NFKD", str(value or "").lower())
        return "".join(ch for ch in text if not unicodedata.combining(ch))
