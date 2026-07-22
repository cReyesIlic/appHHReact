import csv
import gc
import importlib.util
import io
import os
import sqlite3
import tempfile
import unittest
from contextlib import closing
from pathlib import Path
from unittest.mock import AsyncMock, Mock, patch

import pandas as pd
from openpyxl import Workbook

from app.agents.tools import TOOL_SCHEMAS, ToolDispatcher
from app.core.config import settings
from app.rag.parent_child import ParentChildIndexer
from app.services.database_runtime import prepare_runtime_database, runtime_database_status
from app.services.budget_extractor_client import BudgetExtractorClient
from app.services.master_repository import MasterRepository
from app.services.ops_dashboard import OpsDashboardService
from app.services.pipeline_registry import PIPELINE_VERSION, PipelineRegistry, source_signature
from app.services.proposal_sync_service import ProposalSyncService
from app.services.search_filters import SearchFilters
from app.services.sharepoint_client import SharePointClient
from app.services.structured_wiki import StructuredWikiService
from app.services.wiki_auto_compiler import WikiAutoCompiler


class SettingsPathsMixin:
    def patch_settings(self, temp_dir: str):
        base = Path(temp_dir)
        return patch.multiple(
            settings,
            database_dir=str(base / "master.sqlite"),
            master_path=str(base / "master.xlsx"),
            master_path_blob="Master/master.xlsx",
        )


class ParentChildReliabilityTests(SettingsPathsMixin, unittest.TestCase):
    def test_colliding_parent_ids_are_disambiguated_and_old_embeddings_removed(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            indexer = ParentChildIndexer()
            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                conn.execute(
                    """
                    create table rag_child_embeddings (
                        child_id text primary key, parent_id text, codigo text,
                        embedding blob, dim integer, model text, content_hash text
                    )
                    """
                )
                conn.execute(
                    "insert into rag_child_embeddings values (?, ?, ?, ?, ?, ?, ?)",
                    ("old-child", "old-parent", "O-9999", b"old", 1, "model", "hash"),
                )
                conn.commit()

            shared_prefix = "contenido repetido " * 12
            markdown = (
                f"## Fuente\n{shared_prefix}primera variante\n"
                f"## Fuente\n{shared_prefix}segunda variante\n"
            )
            result = indexer.index_parse_result("O-9999", {"text": markdown, "pages": []}, {})

            self.assertEqual(result["parents"], 2)
            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                parent_ids = [row[0] for row in conn.execute(
                    "select parent_id from rag_parent_sections where codigo = ?", ("O-9999",)
                )]
                old_embeddings = conn.execute(
                    "select count(*) from rag_child_embeddings where codigo = ?", ("O-9999",)
                ).fetchone()[0]
            self.assertEqual(len(parent_ids), len(set(parent_ids)))
            self.assertEqual(old_embeddings, 0)

    def test_multiple_documents_keep_their_own_pages_and_sources(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            indexer = ParentChildIndexer()
            result = indexer.index_documents(
                "O-9998",
                [
                    {
                        "parse_result": {
                            "text": "texto tecnico del primer documento " * 8,
                            "pages": [{"pageNumber": 3, "text": "texto tecnico del primer documento " * 8}],
                        },
                        "metadata": {"pdf_name": "primero.pdf", "url": "https://sharepoint/primero.pdf"},
                    },
                    {
                        "parse_result": {
                            "text": "evidencia independiente del segundo archivo " * 8,
                            "pages": [{"pageNumber": 7, "text": "evidencia independiente del segundo archivo " * 8}],
                        },
                        "metadata": {"pdf_name": "segundo.pdf", "url": "https://sharepoint/segundo.pdf"},
                    },
                ],
            )

            self.assertEqual(result["parents"], 2)
            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                rows = conn.execute(
                    "select page_start, metadata from rag_child_chunks where codigo = ? order by page_start",
                    ("O-9998",),
                ).fetchall()
            self.assertEqual([row[0] for row in rows], [3, 7])
            self.assertIn('"pdf_name": "primero.pdf"', rows[0][1])
            self.assertIn('"url": "https://sharepoint/segundo.pdf"', rows[1][1])

    def test_small_overlapping_tail_is_merged_instead_of_duplicated(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            indexer = ParentChildIndexer()
            result = indexer.index_parse_result(
                "O-9997",
                {"text": "x" * 1850, "pages": []},
                {"pdf_name": "documento.pdf"},
            )

            self.assertEqual(result["children"], 2)
            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                lengths = [row[0] for row in conn.execute(
                    "select length(text) from rag_child_chunks where codigo = ? order by json_extract(metadata, '$.child_index')",
                    ("O-9997",),
                )]
            self.assertEqual(lengths, [1000, 1000])

    def test_page_heading_detection_ignores_table_cells(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            indexer = ParentChildIndexer()

            self.assertFalse(indexer._looks_like_heading("x Inmediata x Plazos"))
            self.assertFalse(indexer._looks_like_heading("200 ADistribucion de VDC a control"))
            self.assertTrue(indexer._looks_like_heading("3. Alcance del servicio"))


class ProposalExtractionTests(unittest.TestCase):
    def test_pdf_extraction_preserves_page_numbers(self):
        service = object.__new__(ProposalSyncService)
        pages = [Mock(), Mock()]
        pages[0].extract_text.return_value = "pagina uno"
        pages[1].extract_text.return_value = "pagina dos"
        reader = Mock(pages=pages)

        with patch("PyPDF2.PdfReader", return_value=reader):
            extracted = service._extract_document_any(b"pdf", "pdf", "oferta.pdf")

        self.assertEqual(extracted["text"], "pagina uno\npagina dos")
        self.assertEqual(extracted["pages"], [
            {"pageNumber": 1, "text": "pagina uno"},
            {"pageNumber": 2, "text": "pagina dos"},
        ])


class DatabaseRuntimeTests(SettingsPathsMixin, unittest.TestCase):
    def test_wal_database_is_migrated_to_network_safe_delete_journal(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                self.assertEqual(conn.execute("pragma journal_mode=wal").fetchone()[0], "wal")
                conn.execute("create table sample (id integer primary key)")
                conn.commit()

            result = prepare_runtime_database(attempts=1)

            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                journal = conn.execute("pragma journal_mode").fetchone()[0]
                table = conn.execute(
                    "select count(*) from sqlite_master where type='table' and name='sample'"
                ).fetchone()[0]
            self.assertEqual(result["journal_mode"], "delete")
            self.assertEqual(journal, "delete")
            self.assertEqual(table, 1)
            status = runtime_database_status()
            self.assertTrue(status["network_safe"])
            self.assertGreater(status["size_bytes"], 0)


class QueueReliabilityTests(unittest.TestCase):
    def test_pending_queue_rotates_by_last_attempt(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            manifest = Path(temp_dir) / "sync_manifest.csv"
            with manifest.open("w", encoding="utf-8", newline="") as fh:
                writer = csv.DictWriter(fh, fieldnames=["codigo", "updated_at", "status"])
                writer.writeheader()
                writer.writerow({"codigo": "O-0001", "updated_at": "2026-07-21T12:00:00", "status": "no_files"})
                writer.writerow({"codigo": "O-0003", "updated_at": "2026-07-20T12:00:00", "status": "error"})

            service = object.__new__(ProposalSyncService)
            pending = [{"codigo": "O-0001"}, {"codigo": "O-0002"}, {"codigo": "O-0003"}]
            with patch("app.services.proposal_sync_service.MANIFEST_PATH", str(manifest)):
                ordered = service._order_pending_by_last_attempt(pending)

            self.assertEqual([row["codigo"] for row in ordered], ["O-0002", "O-0003", "O-0001"])

    def test_queue_mix_does_not_starve_new_changed_or_stale(self):
        service = object.__new__(ProposalSyncService)
        mixed = service._mix_queues(
            [
                [{"codigo": "O-0001", "sync_reason": "source_changed"}],
                [{"codigo": "O-0002", "sync_reason": "new_rag"}, {"codigo": "O-0003"}],
                [{"codigo": "O-0004", "sync_reason": "pipeline_stale"}],
            ],
            4,
        )
        self.assertEqual([row["codigo"] for row in mixed], ["O-0001", "O-0002", "O-0004", "O-0003"])

    def test_xlsm_is_extracted_as_excel_content(self):
        workbook = Workbook()
        workbook.active.title = "HH"
        workbook.active.append(["Cargo", "Horas"])
        workbook.active.append(["Ingeniero", 120])
        payload = io.BytesIO()
        workbook.save(payload)

        service = object.__new__(ProposalSyncService)
        text = service._extract_text_any(payload.getvalue(), "xlsm", "oferta.xlsm")
        self.assertIn("## Hoja: HH", text)
        self.assertIn("Ingeniero | 120", text)

    def test_master_winner_codes_are_normalized_and_deduplicated(self):
        service = object.__new__(ProposalSyncService)
        service.master = Mock()
        service.master.all_offers.return_value = [
            {"codigo": "O-274", "estado": "PG", "titulo": "Primera"},
            {"codigo": "O-0274", "estado": "PG", "titulo": "Segunda"},
            {"codigo": "O-2749", "estado": "EP", "titulo": "No ganada"},
        ]

        rows = service._ganadas_master_rows()

        self.assertEqual([row["codigo"] for row in rows], ["O-0274"])
        self.assertEqual(rows[0]["titulo"], "Segunda")

    def test_ops_coverage_joins_unpadded_master_codes_to_canonical_pipeline_codes(self):
        service = object.__new__(OpsDashboardService)
        self.assertEqual(service._code("O-274"), "O-0274")


class SharePointFolderMatchingTests(unittest.IsolatedAsyncioTestCase):
    async def test_manual_sync_normalizes_unpadded_offer_code(self):
        service = object.__new__(ProposalSyncService)
        service.pipeline = Mock()
        service.pipeline.get.return_value = None
        service.sharepoint = Mock()
        service.sharepoint.list_emitido_files = AsyncMock(return_value=[])
        service.sharepoint.list_pdfs = AsyncMock(return_value=[])
        service.sharepoint.select_latest_pdf.return_value = None
        service._record_manifest = Mock()

        result = await service.sync_code("O-274")

        self.assertEqual(result["codigo"], "O-0274")
        service.sharepoint.list_emitido_files.assert_awaited_once_with("O-0274")
        service.pipeline.record_failure.assert_called_once()

    async def test_offer_code_never_falls_back_to_fuzzy_neighbor(self):
        client = object.__new__(SharePointClient)
        client._children = AsyncMock(
            return_value=[
                {"id": "wrong", "name": "O-2370 - Oferta vecina", "folder": {"childCount": 1}},
                {"id": "other", "name": "O-2638 - Otra oferta", "folder": {"childCount": 1}},
            ]
        )

        found = await client._find_child_by_code(Mock(), {}, "drive", "parent", "O-2637")

        self.assertIsNone(found)

    async def test_offer_code_accepts_only_normalized_exact_folder(self):
        client = object.__new__(SharePointClient)
        client._children = AsyncMock(
            return_value=[
                {"id": "correct", "name": "O 2637 - Extension de plazo", "folder": {"childCount": 1}},
                {"id": "wrong", "name": "O-2370 - Oferta vecina", "folder": {"childCount": 1}},
            ]
        )

        found = await client._find_child_by_code(Mock(), {}, "drive", "parent", "O-2637")

        self.assertEqual(found["id"], "correct")

    async def test_offer_code_accepts_explicit_range_but_not_fuzzy(self):
        client = object.__new__(SharePointClient)
        client._children = AsyncMock(
            return_value=[
                {
                    "id": "group",
                    "name": "O-2611 a O-2620 - PDN's 15-21-25-29-32-35-37-38-39-41 - SH-0390 - BHP",
                    "folder": {"childCount": 4},
                },
                {"id": "wrong", "name": "O-2140 - Vecina", "folder": {"childCount": 1}},
            ]
        )

        found = await client._find_child_by_code(Mock(), {}, "drive", "parent", "O-2614")

        self.assertEqual(found["id"], "group")
        self.assertEqual(client._pdn_for_offer_group(found["name"], "O-2614"), 29)
        self.assertIsNone(client._pdn_for_offer_group(found["name"], "O-2621"))

    async def test_default_drive_ignores_preservation_library(self):
        client = object.__new__(SharePointClient)
        client._site_from_url = AsyncMock(return_value={"id": "site"})
        client._request_json = AsyncMock(
            return_value={
                "value": [
                    {
                        "id": "retention",
                        "name": "Biblioteca de suspensiones de conservación",
                        "webUrl": "https://tenant/PreservationHoldLibrary",
                    },
                    {"id": "docs", "name": "Documentos", "webUrl": "https://tenant/Documentos"},
                ]
            }
        )

        drive = await client._default_drive_id(Mock(), {}, "proyectos.vigentes")

        self.assertEqual(drive, "docs")

    def test_project_pdn_document_must_match_filename_or_path(self):
        client = object.__new__(SharePointClient)
        self.assertTrue(
            client._is_exact_pdn_document("SRPS-2630-PC-PDN-21029_P.pdf", "https://tenant/SH-0390", 29)
        )
        self.assertFalse(
            client._is_exact_pdn_document("SRPS-2630-PC-PDN-21091_P.pdf", "https://tenant/SH-0390", 29)
        )

    async def test_registered_mismatched_source_is_removed_when_exact_folder_is_absent(self):
        service = object.__new__(ProposalSyncService)
        service.pipeline = Mock()
        service.pipeline.get.return_value = {
            "source_files": [{"web_url": "https://tenant/01%20Ofertas/O-2370%20-%20Vecina/doc.pdf"}],
            "wiki_entry_id": "wrong-entry",
        }
        service.sharepoint = Mock()
        service.sharepoint.list_emitido_files = AsyncMock(return_value=[])
        service.sharepoint.list_pdfs = AsyncMock(return_value=[])
        service.sharepoint.select_latest_pdf.return_value = None
        service._purge_invalid_index = Mock(return_value={"rag_parent_sections": 1, "wiki_page": True})
        service._record_manifest = Mock()

        result = await service.sync_code("O-2637")

        self.assertEqual(result["status"], "invalid_source_removed")
        service._purge_invalid_index.assert_called_once()
        service.pipeline.record_invalidated.assert_called_once()
        service.sharepoint.download_file.assert_not_called()


class ExcelPipelineTests(SettingsPathsMixin, unittest.IsolatedAsyncioTestCase):
    async def test_excel_assets_are_parsed_into_hh_tables(self):
        workbook = Workbook()
        workbook.active.title = "Estimacion HH"
        workbook.active.append(["Item", "Cargo", "Horas"])
        workbook.active.append(["1.1", "Ingeniero", 120])

        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            path = Path(temp_dir) / "horas.xlsx"
            workbook.save(path)
            service = object.__new__(ProposalSyncService)
            budget = Mock(available=False)
            with patch(
                "app.services.budget_extractor_client.BudgetExtractorClient",
                return_value=budget,
            ):
                result = await service._process_excel_assets(
                    "O-9999",
                    [
                        {
                            "name": "horas.xlsx",
                            "source_file": "horas__item1.xlsx",
                            "kind": "xlsx",
                            "content": path.read_bytes(),
                            "local_path": str(path),
                        }
                    ],
                )

            self.assertEqual(result[0]["hh_status"], "ok")
            self.assertGreater(result[0]["hh_rows"], 0)
            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                rows = conn.execute(
                    "select count(*) from hh_estimate_rows where codigo = 'O-9999'"
                ).fetchone()[0]
            self.assertGreater(rows, 0)

    async def test_duplicate_names_get_distinct_cache_paths(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            base = Path(temp_dir)
            with patch.object(
                type(settings),
                "resolve_path",
                lambda _self, value: base / value,
            ):
                client = object.__new__(SharePointClient)
                first = client.save_pdf_locally("O-9999", "Oferta.xlsx", b"first", item_id="ITEM-001")
                second = client.save_pdf_locally("O-9999", "Oferta.xlsx", b"second", item_id="ITEM-002")

                self.assertNotEqual(first, second)
                self.assertEqual(first.read_bytes(), b"first")
                self.assertEqual(second.read_bytes(), b"second")

    async def test_budget_persistence_converges_cache_alias_to_sharepoint_name(self):
        extracted = {
            "proyecto_filas": [{"descripcion": "Ingeniería", "hh": 10}],
            "tarifas_filas": [],
            "gastos_filas": [{"concepto": "Viaje", "total": 100}],
        }
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            client = BudgetExtractorClient()
            client.persist("O-9999", "Oferta__QRURAKSAHROY.xlsx", extracted)
            client.persist("O-9999", "Oferta.xlsx", extracted)

            stored = client.get_for_codigo("O-9999")

            self.assertEqual(len(stored["filas"]), 1)
            self.assertEqual(len(stored["gastos"]), 1)
            self.assertEqual(stored["filas"][0]["source_file"], "Oferta.xlsx")
            self.assertEqual(len(stored["audit"]), 2)


class PipelineRegistryTests(SettingsPathsMixin, unittest.IsolatedAsyncioTestCase):
    async def test_registry_tracks_files_quality_and_detects_source_change(self):
        initial = [
            {"id": "1", "name": "oferta.pdf", "kind": "pdf", "size": 100, "lastModifiedDateTime": "2026-07-20T10:00:00Z"},
            {"id": "2", "name": "horas.xlsx", "kind": "xlsx", "size": 200, "lastModifiedDateTime": "2026-07-20T11:00:00Z"},
        ]
        changed = [{**initial[0], "size": 150}, initial[1]]
        result = {
            "chunks_parent": 4,
            "chunks_child": 12,
            "embedding_count": 12,
            "wiki_status": "ok",
            "wiki_path": "/tmp/O-9999.md",
            "wiki_entry_id": "entry-1",
            "quality": {"mode": "ai", "rag_score": 88, "wiki_score": 91, "summary": "bien"},
        }
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            registry = PipelineRegistry()
            registry.record_success("O-9999", initial, result)
            state = registry.get("O-9999")
            self.assertEqual(state["pipeline_version"], PIPELINE_VERSION)
            self.assertEqual((state["pdf_count"], state["excel_count"]), (1, 1))
            self.assertEqual((state["rag_quality_score"], state["wiki_quality_score"]), (88, 91))

            service = object.__new__(ProposalSyncService)
            service.pipeline = registry
            service.sharepoint = Mock()
            service.sharepoint.list_emitido_files = AsyncMock(return_value=changed)
            with patch.dict(os.environ, {"SYNC_SOURCE_RECHECK_LIMIT": "10"}, clear=False):
                detected = await service._discover_changed_sources([{"codigo": "O-9999"}])

            self.assertEqual(len(detected), 1)
            self.assertEqual(detected[0]["sync_reason"], "source_changed")
            self.assertEqual(source_signature(detected[0]["_source_files"]), source_signature(changed))
            self.assertEqual(registry.get("O-9999")["source_signature"], source_signature(initial))

            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                conn.execute(
                    "update proposal_pipeline_registry set wiki_pipeline_version = 'wiki-antigua' where codigo = 'O-9999'"
                )
                conn.commit()
            self.assertEqual(registry.status()["needs_reprocess"], 1)

            registry.record_invalidated("O-9999", "fuente asociada a O-1111")
            invalidated = registry.get("O-9999")
            self.assertEqual(invalidated["status"], "invalid_source_removed")
            self.assertEqual((invalidated["pdf_count"], invalidated["excel_count"]), (0, 0))
            self.assertEqual((invalidated["parent_count"], invalidated["embedding_count"]), (0, 0))
            self.assertTrue(invalidated["needs_reprocess"])

    async def test_invalid_recheck_env_falls_back_without_breaking_cycle(self):
        initial = [
            {"id": "1", "name": "oferta.pdf", "size": 100, "lastModifiedDateTime": "2026-07-20T10:00:00Z"}
        ]
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            registry = PipelineRegistry()
            registry.record_success("O-9999", initial, {"wiki_status": "ok", "quality": {}})
            service = object.__new__(ProposalSyncService)
            service.pipeline = registry
            service.sharepoint = Mock()
            service.sharepoint.list_emitido_files = AsyncMock(return_value=initial)
            with patch.dict(
                os.environ,
                {"SYNC_SOURCE_RECHECK_LIMIT": "no-es-numero", "SYNC_SOURCE_RECHECK_CONCURRENCY": "invalido"},
                clear=False,
            ):
                detected = await service._discover_changed_sources([{"codigo": "O-9999"}])
            self.assertEqual(detected, [])

    async def test_wiki_failure_stays_pending_until_backfill_succeeds(self):
        files = [
            {"id": "1", "name": "oferta.pdf", "kind": "pdf", "size": 100,
             "lastModifiedDateTime": "2026-07-20T10:00:00Z"}
        ]
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            registry = PipelineRegistry()
            registry.record_success(
                "O-9999",
                files,
                {
                    "chunks_parent": 2,
                    "chunks_child": 5,
                    "embedding_count": 5,
                    "wiki_status": "error",
                    "wiki_error": "database is locked",
                    "quality": {"mode": "heuristic", "rag_score": 80, "wiki_score": 0},
                },
            )
            pending = registry.get("O-9999")
            self.assertEqual(pending["status"], "partial")
            self.assertTrue(pending["needs_reprocess"])

            registry.record_wiki_success(
                "O-9999",
                {
                    "status": "ok",
                    "path": "/tmp/O-9999.md",
                    "entry_id": "wiki-1",
                    "quality": {"mode": "ai", "rag_score": 89, "wiki_score": 93, "summary": "bien"},
                },
            )
            completed = registry.get("O-9999")
            self.assertEqual(completed["status"], "ok")
            self.assertFalse(completed["needs_reprocess"])
            self.assertEqual((completed["rag_quality_score"], completed["wiki_quality_score"]), (89, 93))

    async def test_wiki_backfill_updates_pipeline_registry(self):
        service = object.__new__(ProposalSyncService)
        service.wiki_compiler = Mock()
        outcome = {
            "codigo": "O-9999", "status": "ok", "path": "/tmp/O-9999.md",
            "entry_id": "wiki-1", "quality": {"rag_score": 88, "wiki_score": 90},
        }
        service.wiki_compiler.compile_for_proposal = AsyncMock(return_value=outcome)
        service.pipeline = Mock()
        service.wiki = Mock()

        result = await service.backfill_wiki(codigos=["O-9999"], force=True, concurrency=1)

        self.assertEqual(result["wiki_ok"], 1)
        service.pipeline.record_wiki_success.assert_called_once_with("O-9999", outcome)
        service.wiki_compiler.compile_for_proposal.assert_awaited_once_with(
            "O-9999", force=True, defer_reindex=True
        )
        service.wiki.reindex_entries.assert_called_once_with()

    async def test_wiki_rebuild_is_resumable_by_schema_marker(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            base = Path(temp_dir)

            def resolve_temp(value):
                path = Path(value)
                return path if path.is_absolute() else base / path

            with patch.object(type(settings), "resolve_path", lambda _self, value: resolve_temp(value)):
                proposals = base / "storage/llm_wiki/proposals"
                proposals.mkdir(parents=True)
                (proposals / "O-1000.md").write_text(
                    "---\nwiki_schema: evidence-v3\n---\n# vigente", encoding="utf-8"
                )
                (proposals / "O-1001.md").write_text(
                    "---\nwiki_schema: legacy\n---\n# antigua", encoding="utf-8"
                )
                service = object.__new__(ProposalSyncService)
                service._codigos_with_rag = Mock(return_value={"O-1000", "O-1001", "O-1002"})

                gaps = service.discover_wiki_gaps()

            self.assertEqual(gaps["wiki_current"], 1)
            self.assertEqual(gaps["wiki_stale"], 1)
            self.assertEqual(gaps["missing_wiki"], 1)
            self.assertEqual(gaps["repair_codes"], ["O-1001", "O-1002"])


class WikiReprocessTests(SettingsPathsMixin, unittest.IsolatedAsyncioTestCase):
    async def test_current_ai_page_is_skipped_without_losing_ai_quality(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            base = Path(temp_dir)
            with patch.object(type(settings), "resolve_path", lambda _self, value: base / value):
                page = base / "storage/llm_wiki/proposals/O-9999.md"
                page.parent.mkdir(parents=True)
                page.write_text(
                    "---\nentry_id: ai-page\nwiki_schema: evidence-v3\nwiki_quality_score: 91\n---\n# vigente",
                    encoding="utf-8",
                )
                compiler = WikiAutoCompiler()
                compiler._collect_rag_material = Mock(side_effect=AssertionError("no debe recompilar"))

                result = await compiler.compile_for_proposal("O-9999")

                self.assertEqual(result["status"], "skipped")
                self.assertEqual(result["entry_id"], "ai-page")
                self.assertEqual(result["quality"]["mode"], "ai")
                self.assertEqual(result["quality"]["wiki_score"], 91)
                del compiler
                gc.collect()

    async def test_legacy_page_is_not_treated_as_current(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            base = Path(temp_dir)
            with patch.object(type(settings), "resolve_path", lambda _self, value: base / value):
                page = base / "storage/llm_wiki/proposals/O-9999.md"
                page.parent.mkdir(parents=True)
                page.write_text("---\nwiki_schema: legacy\n---\n# antigua", encoding="utf-8")
                compiler = WikiAutoCompiler()
                compiler._collect_rag_material = Mock(return_value={"parents": [], "invalid_parent_count": 0})

                result = await compiler.compile_for_proposal("O-9999")

                self.assertEqual(result["status"], "no_rag")
                compiler._collect_rag_material.assert_called_once_with("O-9999")
                del compiler
                gc.collect()

    async def test_reindex_tolerates_duplicate_section_ids(self):
        markdown = """# Wiki\n\n## Grupo\n\n### Repetida\nMismo contenido.\n\n### Repetida\nMismo contenido.\n"""
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            service = StructuredWikiService()

            result = service.build(markdown)

            self.assertEqual(result["sections"], 4)
            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                stored = conn.execute("select count(*) from wiki_sections").fetchone()[0]
            self.assertGreater(stored, 0)

    async def test_forced_reprocess_keeps_one_entry_and_same_file(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            base = Path(temp_dir)

            def resolve_temp(value):
                path = Path(value)
                return path if path.is_absolute() else base / path

            with patch.object(type(settings), "resolve_path", lambda _self, value: resolve_temp(value)):
                ParentChildIndexer().index_parse_result(
                    "O-9999",
                    {
                        "text": "## Alcance\nIngenieria de detalle para sistema de bombeo y entregables verificables.",
                        "pages": [],
                    },
                    {"document_title": "Oferta de prueba"},
                )
                StructuredWikiService.invalidate_sync_cache()
                compiler = WikiAutoCompiler()
                compiler._draft = AsyncMock(return_value={
                    "title": "O-9999 - Oferta de prueba", "category": "propuesta", "tags": ["o-9999"],
                    "content": "## Resumen ejecutivo\nContenido IA [F1].\n\n## Alcance del servicio\nIngeniería de detalle [F1].\n\n## Entregables\nMemoria y planos [F1].\n\n## Evidencia y fuentes\nFuente F1.",
                    "quality_mode": "ai", "rag_quality_score": 85, "wiki_quality_score": 90,
                    "quality_summary": "Ficha IA trazable.", "quality_issues": [],
                })

                first = await compiler.compile_for_proposal("O-9999")
                second = await compiler.compile_for_proposal("O-9999", force=True)

                with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                    count = conn.execute(
                        "select count(*) from wiki_entries where propuestas_referenciadas like '%O-9999%'"
                    ).fetchone()[0]
                    paths = conn.execute(
                        "select distinct file_path from wiki_entries where propuestas_referenciadas like '%O-9999%'"
                    ).fetchall()

                self.assertEqual(first["status"], "ok")
                self.assertEqual(second["status"], "ok")
                self.assertEqual(first["entry_id"], second["entry_id"])
                self.assertEqual(count, 1)
                self.assertEqual(len(paths), 1)
                del compiler
                gc.collect()

    async def test_forced_reprocess_removes_contaminated_duplicate_entries(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            base = Path(temp_dir)

            with patch.object(type(settings), "resolve_path", lambda _self, value: base / value):
                ParentChildIndexer().index_parse_result(
                    "O-9999",
                    {"text": "## Alcance\nIngeniería de detalle válida para O-9999.", "pages": []},
                    {"document_title": "Oferta válida"},
                )
                wiki = StructuredWikiService()
                wiki.upsert_entry(
                    entry_id="legacy-good",
                    title="O-9999 legado",
                    content="Contenido anterior O-9999",
                    source="rag_autocompile",
                    propuestas_referenciadas=["O-9999"],
                    skip_reindex=True,
                )
                wiki.upsert_entry(
                    entry_id="legacy-bad",
                    title="O-9999 contaminado",
                    content="En realidad corresponde a O-8888",
                    source="rag_autocompile",
                    propuestas_referenciadas=["O-9999"],
                    skip_reindex=True,
                )
                proposal_page = base / "storage/llm_wiki/proposals/O-9999.md"
                proposal_page.parent.mkdir(parents=True, exist_ok=True)
                proposal_page.write_text("---\nentry_id: legacy-bad\n---\ncontenido contaminado", encoding="utf-8")

                compiler = WikiAutoCompiler()
                compiler._draft = AsyncMock(return_value={
                    "title": "O-9999 - Oferta válida", "category": "propuesta", "tags": ["o-9999"],
                    "content": "## Resumen ejecutivo\nContenido IA [F1].\n\n## Alcance del servicio\nAlcance válido [F1].\n\n## Entregables\nEntregables válidos [F1].\n\n## Evidencia y fuentes\nFuente F1.",
                    "quality_mode": "ai", "rag_quality_score": 88, "wiki_quality_score": 91,
                    "quality_summary": "Ficha IA trazable.", "quality_issues": [],
                })
                result = await compiler.compile_for_proposal("O-9999", force=True)
                entries = [
                    row for row in wiki.list_entries()
                    if "O-9999" in (row.get("propuestas_referenciadas") or [])
                    and row.get("source") == "rag_autocompile"
                ]

                self.assertEqual(result["status"], "ok")
                self.assertEqual(result["duplicate_cleanup"]["removed"], 2)
                self.assertEqual(len(entries), 1)
                self.assertEqual(entries[0]["id"], result["entry_id"])
                self.assertNotIn("O-8888", entries[0]["content"])
                self.assertTrue(proposal_page.exists())
                del compiler, wiki
                gc.collect()

    async def test_compiler_rejects_rag_whose_file_belongs_to_another_offer(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            base = Path(temp_dir)
            with patch.object(type(settings), "resolve_path", lambda _self, value: base / value):
                ParentChildIndexer().index_parse_result(
                    "O-9999",
                    {
                        "text": "## Alcance\nEste texto fue indexado bajo el código equivocado y no se debe reutilizar.",
                        "pages": [],
                    },
                    {
                        "codigo": "O-9999",
                        "archivo_nombre": "Oferta Técnica O-8888.pdf",
                        "source_path": "storage/proposals/O-8888/Oferta Técnica O-8888.pdf",
                        "document_title": "O-8888 - Oferta contaminante",
                    },
                )
                compiler = WikiAutoCompiler()
                compiler.llm.client = None

                result = await compiler.compile_for_proposal("O-9999", force=True)

                self.assertEqual(result["status"], "invalid_rag")
                self.assertGreater(result["invalid_parent_count"], 0)
                self.assertFalse((base / "storage/llm_wiki/proposals/O-9999.md").exists())
                del compiler
                gc.collect()

    async def test_compiler_does_not_publish_heuristic_wiki_when_ai_is_unavailable(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            base = Path(temp_dir)
            with patch.object(type(settings), "resolve_path", lambda _self, value: base / value):
                ParentChildIndexer().index_parse_result(
                    "O-9999",
                    {"text": "## Alcance\nIngeniería y entregables verificables para la propuesta.", "pages": []},
                    {
                        "codigo": "O-9999",
                        "archivo_nombre": "Oferta Técnica O-9999.pdf",
                        "source_path": "storage/proposals/O-9999/Oferta Técnica O-9999.pdf",
                    },
                )
                compiler = WikiAutoCompiler()
                compiler.llm.client = None

                result = await compiler.compile_for_proposal("O-9999", force=True)

                self.assertEqual(result["status"], "ai_retry")
                self.assertFalse((base / "storage/llm_wiki/proposals/O-9999.md").exists())
                del compiler
                gc.collect()

    def test_material_selection_prioritizes_scope_and_deliverables(self):
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            indexer = ParentChildIndexer()
            metadata = {
                "codigo": "O-9999",
                "archivo_nombre": "Oferta Técnica O-9999.pdf",
                "source_path": "storage/proposals/O-9999/Oferta Técnica O-9999.pdf",
            }
            indexer.index_parse_result(
                "O-9999",
                {
                    "text": (
                        "## Portada\n" + "Texto administrativo sin detalle. " * 20
                        + "\n## Alcance del servicio\n" + "Diseñar la impulsión y verificar transientes. " * 20
                        + "\n## Entregables\n" + "Memoria de cálculo, planos y especificaciones técnicas. " * 20
                    ),
                    "pages": [],
                },
                metadata,
            )
            compiler = WikiAutoCompiler()

            material = compiler._collect_rag_material("O-9999", max_parents=2)

            titles = [row["title"].lower() for row in material["parents"]]
            self.assertTrue(any("alcance" in title for title in titles))
            self.assertTrue(any("entregables" in title for title in titles))
            self.assertEqual(material["invalid_parent_count"], 0)
            del compiler
            gc.collect()


class ChatToolRegistryTests(unittest.TestCase):
    def test_every_exposed_tool_has_exactly_one_dispatch_handler(self):
        schema_names = [item["function"]["name"] for item in TOOL_SCHEMAS]
        dispatcher = ToolDispatcher(ctx=None)
        self.assertEqual(len(schema_names), len(set(schema_names)))
        self.assertEqual(set(schema_names), set(dispatcher._handlers))


class MasterRefreshTests(SettingsPathsMixin, unittest.IsolatedAsyncioTestCase):
    def workbook_bytes(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ofertas"
        for _ in range(21):
            sheet.append([None])
        sheet.append(["Codigo", "Cliente Directo", "Titulo", "Estado"])
        sheet.append(["O-9999", "Cliente prueba", "Oferta prueba", "PG"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    async def test_refresh_uses_sharepoint_validates_and_updates_blob_backup(self):
        content = self.workbook_bytes()
        with tempfile.TemporaryDirectory() as temp_dir, self.patch_settings(temp_dir):
            repository = MasterRepository()
            repository.blob = Mock()
            repository.blob.upload_bytes.return_value = True

            metadata = {
                "name": "master.xlsx",
                "last_modified": "2026-07-20T22:21:29Z",
                "size": len(content),
            }
            with patch(
                "app.services.sharepoint_client.SharePointClient.download_named_file",
                new=AsyncMock(return_value=(content, metadata)),
            ):
                result = await repository.refresh_from_source()

            self.assertEqual(result["source"], "sharepoint")
            self.assertEqual(result["rows_loaded"], 1)
            self.assertTrue(result["blob_updated"])
            self.assertTrue(Path(settings.master_path).exists())
            repository.blob.upload_bytes.assert_called_once()
            with closing(sqlite3.connect(settings.sqlite_path)) as conn:
                row = conn.execute("select codigo, estado from oferta").fetchone()
            self.assertEqual(row, ("O-9999", "PG"))


class MasterCodeFilterTests(unittest.TestCase):
    def test_offer_code_filter_is_exact_not_substring(self):
        frame = pd.DataFrame(
            [
                {"codigo": "O-0274", "cod_proy": "10.0"},
                {"codigo": "O-2749", "cod_proy": "274.0"},
                {"codigo": "O-1274", "cod_proy": "20.0"},
            ]
        )
        repo = MasterRepository()

        mask = repo._filter_mask(frame, SearchFilters(codigos=["O-274"]))

        self.assertEqual(frame[mask]["codigo"].tolist(), ["O-0274"])

    def test_project_code_filter_matches_numeric_project_column_exactly(self):
        frame = pd.DataFrame(
            [
                {"codigo": "O-2000", "cod_proy": "390.0"},
                {"codigo": "O-2001", "cod_proy": "1390.0"},
            ]
        )
        repo = MasterRepository()
        mask = repo._filter_mask(frame, SearchFilters(codigos=["SH-0390"]))

        self.assertEqual(frame[mask]["codigo"].tolist(), ["O-2000"])


class SchedulerReliabilityTests(unittest.IsolatedAsyncioTestCase):
    @unittest.skipUnless(importlib.util.find_spec("apscheduler"), "APScheduler no instalado en el Python local")
    async def test_scheduler_runs_once_daily_in_chile_timezone(self):
        from app.services import scheduler

        scheduler.shutdown_scheduler()
        with patch.dict(
            os.environ,
            {
                "SYNC_SCHEDULE_ENABLED": "true",
                "SYNC_SCHEDULE_HOURS": "2",
                "SYNC_SCHEDULE_TZ": "America/Santiago",
                "SYNC_SCHEDULE_MINUTE": "15",
                "WIKI_REBUILD_ON_STARTUP": "false",
            },
            clear=False,
        ):
            info = scheduler.start_scheduler()
            status = scheduler.scheduler_status()
            scheduler.shutdown_scheduler()

        self.assertEqual(info["runs_per_day"], 1)
        self.assertEqual(info["hours"], [2])
        self.assertEqual(status["timezone"], "America/Santiago")
        self.assertEqual(
            {job["id"] for job in status["jobs"]},
            {"sync_ganadas_periodic", "storage_monitor_daily"},
        )
        self.assertTrue(all(job["timezone"] == "America/Santiago" for job in status["jobs"]))
        sync_job = next(job for job in status["jobs"] if job["id"] == "sync_ganadas_periodic")
        self.assertIn("hour='2'", sync_job["trigger"])


if __name__ == "__main__":
    unittest.main()
